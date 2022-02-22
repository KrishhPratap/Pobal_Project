import time
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
import openpyxl as op


class Signup:
    # Locator of all the element on Landing
    Name_xpath = "//*[@id='Name']"
    Email_xpath = "//*[@id='Email']"
    Country_dropdown_xpath = "//div[@class='iti__selected-flag dropdown-toggle']"
    country_name = "//span[@class='iti__country-name']"
    PhoneNo_xpath = "//input[@id='adsasd']"
    Password_xpath = "//*[@id='Password']"
    PassEye_xpath = "/html/body/app-root/app-auth/div/div/div/div[2]/app-sign-up/div/div/div[2]/form/div[4]/em"
    Confirm_Password_xpath = "//*[@id='confirmPass']"
    Confirm_eye_xpath = "/html/body/app-root/app-auth/div/div/div/div[2]/app-sign-up/div/div/div[2]/form/div[5]/em"
    Terms_condition_checkbox_xpath = "//input[@type='checkbox']"
    SignUp_btn_xpath = "//div[@class='btnDiv']"
    SignIn_btn_xpath = "//span[@class='sign-text']"

    def __init__(self, driver):
        self.driver = driver

    """
    def setName(self, username):
        self.driver.find_element(By.XPATH, self.Name_xpath).send_keys(username)


    def setEmailId(self, emailaddress):
        self.driver.find_element(By.XPATH, self.Email_xpath).send_keys(emailaddress)

    def setCountry(self, countryname):
        self.driver.find_element(By.XPATH, self.Country_dropdown_xpath).click()
        cname = self.driver.find_elements(By.XPATH, self.country_name)
        # print(len(cname))
        for i in cname:
            # print(i.text)
            if i.text == countryname:
                i.click()
                break

    def setPhoneNo(self, phonenumber):
        self.driver.find_element(By.XPATH, self.PhoneNo_xpath).send_keys(phonenumber)

    def setPassword(self, password):
        self.driver.find_element(By.XPATH, self.Password_xpath).send_keys(password)

    def pass_eyeBtn(self):
        self.driver.find_element(By.XPATH, self.PassEye_xpath).click()

    def confirm_Pass_eyeBtn(self):
        self.driver.find_element(By.XPATH, self.Confirm_eye_xpath).click()

    def set_confirm_password(self, confirm_pass):
        self.driver.find_element(By.XPATH, self.Confirm_Password_xpath).send_keys(confirm_pass)

    def terms_condition(self):
        self.driver.find_element(By.XPATH, self.Terms_condition_checkbox_xpath).click()

    def signupbtn(self):
        self.driver.find_element(By.XPATH, self.SignUp_btn_xpath).click()

    def signInbtn(self):
        self.driver.find_element(By.XPATH, self.SignIn_btn_xpath).click()

        """

    def signup(self):
        excel_file = "/home/algoworks/Documents/pobalcreds.xlsx"
        excel_worksheet = "Sheet1"
        wb = op.load_workbook(excel_file)  # workbook
        ws = wb[excel_worksheet]  # name of excelfile and pile path
        row_num = ws.max_row  # variable for work sheet
        col_num = ws.max_column
        print("the no. of rows is", row_num, "and the no. of column is ", col_num)
        user_name = self.driver.find_element(By.XPATH, self.Name_xpath)
        email_address = self.driver.find_element(By.XPATH, self.Email_xpath)
        country_drop = self.driver.find_element(By.XPATH, self.Country_dropdown_xpath)

        phone_no = self.driver.find_element(By.XPATH, self.PhoneNo_xpath)
        password = self.driver.find_element(By.XPATH, self.Password_xpath)
        confirmpass = self.driver.find_element(By.XPATH, self.Confirm_Password_xpath)

        for row in range(2, ws.max_row + 1):
            username = str(ws.cell(row, 1).value)
            user_name.clear()
            user_name.send_keys(username)
            time.sleep(3)

            email = str(ws.cell(row, 2).value)
            email_address.send_keys(Keys.CONTROL, 'a')
            email_address.send_keys(Keys.BACKSPACE)
            email_address.send_keys(email)
            time.sleep(3)

            country_drop.click()
            country = str(ws.cell(row, 3).value)
            print(country)
            cname = self.driver.find_elements(By.XPATH, self.country_name)
            print(len(cname))
            for i in cname:
                print(i.text)
                if i.text == country:
                    print(i.text)
                    i.click()
                    break
                else:
                    print("Not working")
            time.sleep(1)

            phone = str(ws.cell(row, 4).value)
            phone_no.clear()
            phone_no.send_keys(phone)
            time.sleep(3)

            paswd = str(ws.cell(row, 5).value)
            password.clear()
            password.send_keys(paswd)
            time.sleep(3)

            confpaswd = str(ws.cell(row, 6).value)
            confirmpass.clear()
            confirmpass.send_keys(confpaswd)
            time.sleep(3)

        self.driver.find_element(By.XPATH, self.Terms_condition_checkbox_xpath).click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.SignUp_btn_xpath).click()
        self.driver.find_element(By.XPATH, self.SignIn_btn_xpath).click()
